from threading import Thread
import csv
from time import sleep
import pandas as pd
from openpyxl import load_workbook
import yaml


class FormatThread(Thread):
    def __init__(self, threadID, name):
        Thread.__init__(self)
        self.name = name
        self.thread_id = threadID
        self.file_name = ""
        self.status = "Starting"
        self.output_file = ""

    def run(self):
        self.output_file = self.file_name.rsplit('.', 1)[0] + '.csv'
        self.status = "Getting Settings"
        # format_file(self.file_name, 'config.yml', self.file_name.rsplit('.', 1)[0] + '.csv')
        with open('config.yml') as settings_file:
            settings = yaml.load(settings_file)
            self.status = "Arranging Headers"
        print(self.status)
        self.status = formatter(self.file_name, settings, 'intermediate.csv')
        print(self.status)
        self.status = output(stacker('intermediate.csv', settings), 'intermediate.csv')
        print(self.status)
        self.status = final_formatting('intermediate.csv', self.output_file)
        print(self.status)

def formatter(inbound_path, settings, outbound_path):
    """Reads the excel sheet and formats it for later transformations.
    Saves to intermediate.csv"""
    workbook = load_workbook(inbound_path, read_only=True)
    sheet = workbook.active
    with open(outbound_path, 'w') as intermediate:
        writer = csv.writer(intermediate)
        for row_num, row in enumerate(sheet.rows):
            if row_num == settings['Type Row'] - 1:
                new_row = []
                for col_num, cell in enumerate(row):
                    if col_num >= settings['First Store Column'] - 1:
                        new_row.append(combine_store_and_type(sheet, cell, col_num, settings))
                    else:
                        new_row.append(cell.value)
                row = new_row
            else:
                row = [c.value for c in row]
            writer.writerow(row)
    return "Rotating"

def combine_store_and_type(sheet, cell, col_num, settings):
    store = sheet.cell(row=settings['Store Row'], column=(
        (col_num - (settings['First Store Column'] - 1))
        // settings['Number of Types'] * settings['Number of Types']
        + settings['First Store Column']
    )).value
    return str(cell.value) + ',' + store

def stacker(path, settings):
    # print("Stacking data")
    report = pd.read_csv(
        path,
        header=settings["Type Row"] - 1,
        index_col=settings["Index Columns"],
    )
    return report.stack()

def output(report, path):
    # print("Saving workbook to " + path)
    report.to_frame(name="Values").to_csv(path)
    return "Cleaning Up"

def format_file(inbound_path, settings_path, outbound_path):
    print("Getting settings")
    with open(settings_path) as settings_file:
        settings = yaml.load(settings_file)
    formatter(inbound_path, settings, 'intermediate.csv')
    output(stacker('intermediate.csv', settings), 'intermediate.csv')
    final_formatting('intermediate.csv', outbound_path)
    print('Format complete. Saved to ' + outbound_path)

def final_formatting(inbound_path, outbound_path):
    with open(inbound_path, 'r') as inbound_file:
        with open(outbound_path, 'w') as outbound_file:
            reader = csv.reader(inbound_file)
            writer = csv.writer(outbound_file)
            for index, row in enumerate(reader):
                if index == 0:
                    writer.writerow(['Dept', 'Style', 'Type', 'Store', 'Values'])
                else:
                    writer.writerow(row[:2] + row[2].split(',') + row[3:])
    return "Complete"

if __name__ == '__main__':
    with open("config.yml") as file:
        CONFIG = yaml.load(file)
    formatter("uploads/Marco_Bicego_Style_by_Store_Selling_Week_November_Week_1.xlsx", CONFIG,
              'intermediate.csv')
    output(stacker('intermediate.csv', CONFIG), 'intermediate.csv')
    final_formatting('intermediate.csv', 'Test2.csv')
